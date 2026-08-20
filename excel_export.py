"""
Экспорт данных в Excel.
"""

import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import load_config
from constants import (
    INSTITUTION_LINES, APPROVER_TITLE, APPROVER_NAME, UMO_NAME,
    TABLE_HEADERS_ROW1, TABLE_HEADERS_ROW2, COL_WIDTHS,
    ALL_KAFEDRY, ALL_GODY,
)
from database import fetch_nagruzka, fetch_teacher_info, fetch_kafedra_i_fakultet_info


# =========================================================
#  УТИЛИТЫ
# =========================================================

def transliterate_to_latin(text: str) -> str:
    if not text:
        return ""
    translit_map = {
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'E',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
        'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    }
    result = []
    for char in text:
        if char in translit_map:
            result.append(translit_map[char])
        elif char == ' ' or char == '.':
            result.append('_')
        elif char.isalnum():
            result.append(char)
    transliterated = ''.join(result)
    transliterated = re.sub(r'_+', '_', transliterated)
    transliterated = transliterated.strip('_')
    return transliterated or "unknown"


def shorten_fio(fio: str) -> str:
    if not fio:
        return ""
    fio = fio.strip()
    if not fio:
        return ""
    if re.match(r'^[А-ЯЁ][а-яё]+(\s+[А-ЯЁ]\.[А-ЯЁ]\.)?$', fio):
        return fio
    parts = re.split(r'\s+', fio)
    if len(parts) == 1:
        parts = re.findall(r'[А-ЯЁ][а-яё]+', fio)
    if len(parts) == 0:
        return fio
    surname = parts[0]
    if len(parts) >= 3:
        return f"{surname} {parts[1][0].upper()}.{parts[2][0].upper()}."
    elif len(parts) == 2:
        return f"{surname} {parts[1][0].upper()}."
    return surname


def _safe_sheet_name(name: str, used: set):
    clean = re.sub(r'[\\/*?:\[\]]', "", name)[:31] or "Лист"
    base = clean
    i = 2
    while clean in used:
        suffix = f" ({i})"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


def _get_template_settings() -> dict:
    """
    Возвращает настройки шаблона из config.json.
    Если в конфиге значение пустое — использует значение по умолчанию из constants.py.
    """
    cfg = load_config()

    # Шапка учреждения
    inst_raw = cfg.get("template_institution_lines", "")
    if inst_raw and inst_raw.strip():
        inst_lines = [line.strip() for line in inst_raw.split("\n") if line.strip()]
    else:
        inst_lines = list(INSTITUTION_LINES)

    # Должность утверждающего
    approver_title = cfg.get("template_approver_title", "").strip()
    if not approver_title:
        approver_title = APPROVER_TITLE

    # ФИО утверждающего
    approver_name = cfg.get("template_approver_name", "").strip()
    if not approver_name:
        approver_name = APPROVER_NAME

    # ФИО начальника УМО
    umo_name = cfg.get("template_umo_name", "").strip()
    if not umo_name:
        umo_name = UMO_NAME

    return {
        "inst_lines": inst_lines,
        "approver_title": approver_title,
        "approver_name": approver_name,
        "umo_name": umo_name,
    }


# =========================================================
#  ФОРМИРОВАНИЕ ЛИСТА
# =========================================================

def write_teacher_sheet(ws, teacher_info, kafedra_name, god_label, columns, rows,
                        mode="semester", kaf_info=None, numbering="original",
                        filter_note=None):
    if kaf_info is None:
        kaf_info = {"Заведующий": "", "Декан": ""}

    # Читаем актуальные настройки шаблона
    tpl = _get_template_settings()
    inst_lines = tpl["inst_lines"]
    approver_title = tpl["approver_title"]
    approver_name = tpl["approver_name"]
    umo_name = tpl["umo_name"]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    n_cols = len(TABLE_HEADERS_ROW1)
    for idx, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ---------- Шапка учреждения ----------
    r = 1
    for line in inst_lines:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols - 3)
        cell = ws.cell(row=r, column=1, value=line)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1

    # ---------- Блок "УТВЕРЖДАЮ" ----------
    approve_start = r - len(inst_lines)
    ws.merge_cells(start_row=approve_start, start_column=n_cols - 2,
                   end_row=approve_start + 1, end_column=n_cols)
    approve_cell = ws.cell(
        row=approve_start, column=n_cols - 2,
        value=(
            f"УТВЕРЖДАЮ:\n"
            f"{approver_title}\n"
            f"_____________ {approver_name}\n"
            f"«___»___________ {datetime.now().year} г."
        )
    )
    approve_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    approve_cell.font = Font(size=9)

    # ---------- Заголовок отчёта ----------
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    title_cell = ws.cell(row=r, column=1, value=f"УЧЕБНАЯ НАГРУЗКА НА {god_label} УЧЕБНЫЙ ГОД")
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")
    r += 1

    if filter_note:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        note_cell = ws.cell(row=r, column=1, value=f"Применены фильтры: {filter_note}")
        note_cell.font = Font(italic=True, size=8, color="888888")
        note_cell.alignment = Alignment(horizontal="center")
        r += 1

    r += 1

    # ---------- Информация о преподавателе ----------
    for label, value in [
        ("Кафедра:", kafedra_name or ""),
        ("ФИО преподавателя:", teacher_info.get("ФИО", "")),
        ("Должность:", teacher_info.get("Должность", "")),
        ("Условия работы:", teacher_info.get("Условия", "")),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1

    # ---------- Вспомогательные функции ----------
    col_map = {name: i for i, name in enumerate(columns)}

    def val(row, key):
        return row[col_map[key]] if key in col_map else None

    def write_table_header(start_row):
        header_row1 = start_row
        header_row2 = start_row + 1
        for col_idx, text in enumerate(TABLE_HEADERS_ROW1, start=1):
            cell = ws.cell(row=header_row1, column=col_idx, value=text if text else None)
            cell.font = Font(bold=True, size=9)
            cell.alignment = center
            cell.border = border
        for col_idx, text in enumerate(TABLE_HEADERS_ROW2, start=1):
            cell = ws.cell(row=header_row2, column=col_idx, value=text if text else None)
            cell.font = Font(bold=True, size=9)
            cell.alignment = center
            cell.border = border
        for c in [1, 2, 3, 4, 5, 6, 7, 8]:
            ws.merge_cells(start_row=header_row1, start_column=c, end_row=header_row2, end_column=c)
        ws.merge_cells(start_row=header_row1, start_column=9, end_row=header_row1, end_column=10)
        ws.merge_cells(start_row=header_row1, start_column=11, end_row=header_row1, end_column=13)
        ws.merge_cells(start_row=header_row1, start_column=14, end_row=header_row1, end_column=15)
        return header_row2 + 1

    def write_block(start_row, title, block_rows, total_label=None):
        if total_label is None:
            total_label = f"Итого за {title.lower()}:"
        rr = start_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=n_cols)
        title_cell = ws.cell(row=rr, column=1, value=title)
        title_cell.font = Font(bold=True, size=10, italic=True)
        rr += 1
        rr = write_table_header(rr)

        s_aud = s_dr = s_itogo = s_byudzhet = s_vneb = 0.0
        row_counter = 1

        if not block_rows:
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=n_cols)
            empty_cell = ws.cell(row=rr, column=1, value="(нет нагрузки за этот период)")
            empty_cell.font = Font(italic=True, color="888888", size=9)
            rr += 1
        else:
            for row in block_rows:
                kurs = val(row, "Курс")
                sem = val(row, "Семестр")
                kurs_sem = "/".join(str(x) for x in (kurs, sem) if x is not None)

                aud = val(row, "НагрузкаАуд") or 0
                dr = val(row, "НагрузкаДр") or 0
                itogo = val(row, "Итого") or 0
                byudzhet = val(row, "Бюджет") or 0
                vneb = val(row, "Внебюджет") or 0

                s_aud += float(aud)
                s_dr += float(dr)
                s_itogo += float(itogo)
                s_byudzhet += float(byudzhet)
                s_vneb += float(vneb)

                line_number = row_counter if numbering == "sequential" else val(row, "НомерСтроки")
                if numbering == "sequential":
                    row_counter += 1

                values = [
                    line_number, val(row, "Группа"), val(row, "Блок"),
                    val(row, "Дисциплина"), kurs_sem, val(row, "ВидЗанятий"),
                    val(row, "ВидКонтроля"), val(row, "Студентов"), val(row, "НомерПотока"),
                    val(row, "ИндикаторПотока"), aud, dr, itogo, byudzhet, vneb,
                ]
                for col_idx, v in enumerate(values, start=1):
                    cell = ws.cell(row=rr, column=col_idx, value=v)
                    cell.border = border
                    cell.alignment = center if col_idx != 4 else left
                    cell.font = Font(size=9)
                rr += 1

        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=10)
        itogo_label = ws.cell(row=rr, column=1, value=total_label)
        itogo_label.font = Font(bold=True)
        itogo_label.alignment = Alignment(horizontal="right")
        for col_idx, total in zip([11, 12, 13, 14, 15], [s_aud, s_dr, s_itogo, s_byudzhet, s_vneb]):
            cell = ws.cell(row=rr, column=col_idx, value=round(total, 2))
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = center
        for col_idx in range(1, 11):
            ws.cell(row=rr, column=col_idx).border = border
        rr += 2
        return rr, (s_aud, s_dr, s_itogo, s_byudzhet, s_vneb)

    # ---------- Разделение строк по семестрам ----------
    autumn_rows  = [row for row in rows if val(row, "Семестр") is not None and int(val(row, "Семестр")) % 2 == 1]
    spring_rows  = [row for row in rows if val(row, "Семестр") is not None and int(val(row, "Семестр")) % 2 == 0]
    unknown_rows = [row for row in rows if val(row, "Семестр") is None]

    if mode == "semester":
        r, sums_autumn = write_block(r, "Осенний семестр", autumn_rows)
        r, sums_spring = write_block(r, "Весенний семестр", spring_rows)
        sums_unknown = (0, 0, 0, 0, 0)
        if unknown_rows:
            r, sums_unknown = write_block(r, "Без указания семестра", unknown_rows)
        grand = [a + b + c for a, b, c in zip(sums_autumn, sums_spring, sums_unknown)]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        grand_label = ws.cell(row=r, column=1, value="ИТОГО ЗА ГОД:")
        grand_label.font = Font(bold=True, size=11)
        grand_label.alignment = Alignment(horizontal="right")
        for col_idx, total in zip([11, 12, 13, 14, 15], grand):
            cell = ws.cell(row=r, column=col_idx, value=round(total, 2))
            cell.font = Font(bold=True, size=11)
            cell.border = border
            cell.alignment = center
        for col_idx in range(1, 11):
            ws.cell(row=r, column=col_idx).border = border
        r += 2
    elif mode == "autumn":
        r, _ = write_block(r, "Осенний семестр", autumn_rows, total_label="Итого за осенний семестр:")
    elif mode == "spring":
        r, _ = write_block(r, "Весенний семестр", spring_rows, total_label="Итого за весенний семестр:")
    else:
        r, _ = write_block(r, "Вся нагрузка за год", list(rows), total_label="Итого за год:")

    # ---------- Подписи ----------
    ws.cell(row=r, column=1, value="Преподаватель _________________")
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    ws.cell(row=r, column=4, value=teacher_info.get("ФИО", ""))
    r += 2

    ws.cell(row=r, column=1, value="СОГЛАСОВАНО:")
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 2

    zav_fio = shorten_fio(kaf_info.get("Заведующий", ""))
    zav_text = f"Заведующий кафедрой _____________ {zav_fio}" if zav_fio else "Заведующий кафедрой _____________"
    ws.cell(row=r, column=1, value=zav_text)

    dek_fio = shorten_fio(kaf_info.get("Декан", ""))
    dek_text = f"Декан факультета _____________ {dek_fio}" if dek_fio else "Декан факультета _____________"
    ws.cell(row=r, column=5, value=dek_text)

    # ФИО начальника УМО теперь берётся из конфига
    ws.cell(row=r, column=9, value=f"Начальник УМО _____________ {umo_name}")


# =========================================================
#  СБОРКА КНИГИ
# =========================================================

def build_workbook(conn, teachers, kafedra_filter, god_label, mode="semester", numbering="original",
                    vidy_zanyatiy=None, vidy_kontrolya=None, finance=None):
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    kaf_info_cache = {}

    note_parts = []
    if vidy_zanyatiy:
        note_parts.append("вид занятий: " + ", ".join(vidy_zanyatiy))
    if vidy_kontrolya:
        note_parts.append("вид контроля: " + ", ".join(vidy_kontrolya))
    if finance == "budget":
        note_parts.append("только бюджет")
    elif finance == "vneb":
        note_parts.append("только внебюджет")
    filter_note = "; ".join(note_parts) if note_parts else None

    written_any = False
    for kod, fio in teachers:
        columns, rows = fetch_nagruzka(conn, kod, god_label, vidy_zanyatiy, vidy_kontrolya, finance)
        info = fetch_teacher_info(conn, kod)

        kafedra_name = kafedra_filter if (kafedra_filter and kafedra_filter != ALL_KAFEDRY) else None
        if not kafedra_name and rows:
            col_map = {name: i for i, name in enumerate(columns)}
            if "Кафедра" in col_map:
                kafedra_name = rows[0][col_map["Кафедра"]]

        if kafedra_name not in kaf_info_cache:
            kaf_info_cache[kafedra_name] = fetch_kafedra_i_fakultet_info(conn, kafedra_name) if kafedra_name else {"Заведующий": "", "Декан": ""}
        kaf_info = kaf_info_cache[kafedra_name]

        sheet_name = _safe_sheet_name(fio or f"Преп_{kod}", used_names)
        ws = wb.create_sheet(sheet_name)
        god_display = god_label if (god_label and god_label != ALL_GODY) else "____"
        write_teacher_sheet(ws, info, kafedra_name, god_display, columns, rows, mode, kaf_info, numbering,
                            filter_note)
        written_any = True

    if not written_any:
        wb.create_sheet("Преподаватели")
    return wb